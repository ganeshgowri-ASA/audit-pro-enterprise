"""
Excel Export Utilities
AuditPro Enterprise - Export data to Excel
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os
from config.settings import BASE_DIR


def export_to_excel(data: pd.DataFrame, filename: str, sheet_name: str = "Data") -> str:
    """
    Export DataFrame to Excel with basic formatting

    Args:
        data: DataFrame to export
        filename: Output filename
        sheet_name: Sheet name

    Returns:
        str: Path to generated Excel file
    """
    output_path = BASE_DIR / "data" / "exports" / filename
    os.makedirs(output_path.parent, exist_ok=True)

    # Write to Excel
    with pd.ExcelWriter(str(output_path), engine='openpyxl') as writer:
        data.to_excel(writer, sheet_name=sheet_name, index=False)

    return str(output_path)


def export_audit_data(audits_df: pd.DataFrame, filename: str = None) -> str:
    """
    Export audit data to formatted Excel

    Args:
        audits_df: DataFrame containing audit data
        filename: Output filename (optional)

    Returns:
        str: Path to generated Excel file
    """
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"audit_export_{timestamp}.xlsx"

    output_path = BASE_DIR / "data" / "exports" / filename
    os.makedirs(output_path.parent, exist_ok=True)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Audits"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1f77b4", end_color="1f77b4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write data
    for r_idx, row in enumerate(dataframe_to_rows(audits_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border

            # Format header row
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save workbook
    wb.save(str(output_path))

    return str(output_path)


def export_nc_ofi_data(nc_df: pd.DataFrame, filename: str = None) -> str:
    """
    Export NC/OFI data to formatted Excel

    Args:
        nc_df: DataFrame containing NC/OFI data
        filename: Output filename (optional)

    Returns:
        str: Path to generated Excel file
    """
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"nc_ofi_export_{timestamp}.xlsx"

    output_path = BASE_DIR / "data" / "exports" / filename
    os.makedirs(output_path.parent, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "NC_OFI"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="d62728", end_color="d62728", fill_type="solid")

    # Write data with formatting
    for r_idx, row in enumerate(dataframe_to_rows(nc_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Auto-adjust columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    wb.save(str(output_path))
    return str(output_path)


def export_multi_sheet_report(data_dict: dict, filename: str = None) -> str:
    """
    Export multiple DataFrames to multi-sheet Excel

    Args:
        data_dict: Dictionary with sheet_name: DataFrame pairs
        filename: Output filename (optional)

    Returns:
        str: Path to generated Excel file
    """
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"multi_report_{timestamp}.xlsx"

    output_path = BASE_DIR / "data" / "exports" / filename
    os.makedirs(output_path.parent, exist_ok=True)

    with pd.ExcelWriter(str(output_path), engine='openpyxl') as writer:
        for sheet_name, df in data_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    return str(output_path)


def create_summary_dashboard(summary_data: dict, filename: str = None) -> str:
    """
    Create Excel dashboard with summary statistics

    Args:
        summary_data: Dictionary containing summary statistics
        filename: Output filename (optional)

    Returns:
        str: Path to generated Excel file
    """
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"dashboard_{timestamp}.xlsx"

    output_path = BASE_DIR / "data" / "exports" / filename
    os.makedirs(output_path.parent, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"

    # Title
    ws['A1'] = "AuditPro Enterprise Dashboard"
    ws['A1'].font = Font(size=16, bold=True, color="1f77b4")
    ws['A1'].alignment = Alignment(horizontal="center")
    ws.merge_cells('A1:D1')

    # Date
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].alignment = Alignment(horizontal="center")
    ws.merge_cells('A2:D2')

    # Summary data
    row = 4
    for key, value in summary_data.items():
        ws[f'A{row}'] = key
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        row += 1

    # Formatting
    for col in ['A', 'B']:
        ws.column_dimensions[col].width = 30

    wb.save(str(output_path))
    return str(output_path)
